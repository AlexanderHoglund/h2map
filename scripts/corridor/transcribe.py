"""Green Corridor workbook transcription (Phase 0).

Reads Green_Corridor_Model_Simplified_30_07.xlsx twice (formulas + cached
values) and deterministically emits:

  docs/excel-transcription-dump.json
      Every non-empty cell of all 9 tabs: {sheet, cell, formula, value}.
      The machine-readable companion to docs/excel-transcription.md.

  fixtures/golden/corridor/excel-baseline.expected.json
      The golden EXPECTED values, projected from the workbook's CACHED
      values (data_only=True). The expected side of the fixture comes from
      the workbook itself, never from the engine — which is why there is no
      "update goldens" script for the corridor engine. Never re-evaluate
      formulas here (an evaluator would introduce genuine drift); cached
      values only.

Also prints the workbook sha256 (recorded in the transcription doc; any
workbook change requires a NEW bundle id + NEW fixture file, never edits).

Usage:  python scripts/corridor/transcribe.py
Deterministic: re-running on the same workbook produces byte-identical files.
"""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "Green_Corridor_Model_Simplified_30_07.xlsx"
DUMP_OUT = ROOT / "docs" / "excel-transcription-dump.json"
EXPECTED_OUT = ROOT / "fixtures" / "golden" / "corridor" / "excel-baseline.expected.json"

SHEETS = [
    "Cover", "Cargo", "Vessel", "Fuel", "Port",
    "Regulation", "Calculation", "Output", "Data_tables",
]

# Fixture horizon: Cargo!D15 = 20 model years, Calculation columns D..W.
HORIZON_YEARS = 20
FIRST_YEAR_COL = 4  # column D


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def cell_name(row: int, col: int) -> str:
    return f"{openpyxl.utils.get_column_letter(col)}{row}"


def build_dump(wb_f, wb_v) -> dict:
    sheets = {}
    for name in SHEETS:
        wsf, wsv = wb_f[name], wb_v[name]
        cells = []
        for row in range(1, wsf.max_row + 1):
            for col in range(1, wsf.max_column + 1):
                f = wsf.cell(row=row, column=col).value
                v = wsv.cell(row=row, column=col).value
                if f is None and v is None:
                    continue
                entry = {"cell": cell_name(row, col)}
                if isinstance(f, str) and f.startswith("="):
                    entry["formula"] = f
                    entry["value"] = v
                else:
                    entry["value"] = f if f is not None else v
                cells.append(entry)
        sheets[name] = cells
    return {"workbook": WORKBOOK.name, "sha256": sha256(WORKBOOK), "sheets": sheets}


def year_row(ws, row: int) -> list:
    """The 20 horizon-year cached values of one Calculation row (cols D..W)."""
    out = []
    for i in range(HORIZON_YEARS):
        v = ws.cell(row=row, column=FIRST_YEAR_COL + i).value
        out.append(0 if v is None else v)
    return out


def build_expected(wb_v) -> dict:
    calc = wb_v["Calculation"]
    fuel = wb_v["Fuel"]
    vessel = wb_v["Vessel"]
    output = wb_v["Output"]

    def c(row: int, col: int = FIRST_YEAR_COL):
        return calc.cell(row=row, column=col).value

    summary = {
        "greenTotalPvUsdM": c(70),
        "fossilTotalPvUsdM": c(71),
        "gapPvUsdM": c(79),
        "etsGreenPvUsdM": c(72),
        "fuelEuGreenPvUsdM": c(73),
        "ira45zGreenPvUsdM": c(74),
        "selfDesignedGreenPvUsdM": c(75),
        "etsFossilPvUsdM": c(76),
        "fuelEuFossilPvUsdM": c(77),
        "selfDesignedFossilPvUsdM": c(78),
        "cargoUnitsLifetime": c(80),
        "co2AbatedTonnes": c(81),
        "greenCapexPvUsdM": c(82),
        "greenOpexPvUsdM": c(83),
        "fossilCapexPvUsdM": c(84),
        "fossilOpexPvUsdM": c(85),
        # Output!D26 / D31 — derived unit metrics (gap ×1e6 ÷ units / ÷ tCO2).
        "costPerUnitUsd": output.cell(row=26, column=4).value,
        "costPerTonneCo2Usd": output.cell(row=31, column=4).value,
    }

    intermediates = {
        # Resolved inputs the engine must reproduce through its resolution layer.
        "greenFuelTonnesPerVesselYear": fuel.cell(row=15, column=5).value,   # Fuel!E15
        "fossilFuelTonnesPerVesselYear": fuel.cell(row=28, column=5).value,  # Fuel!E28
        "greenVesselCapexUsdM": vessel.cell(row=12, column=5).value,         # Vessel!E12
    }

    green = {
        "totalCapexUsdM": year_row(calc, 25),
        "totalOpexUsdM": year_row(calc, 26),
        "etsUsdM": year_row(calc, 28),
        "fuelEuUsdM": year_row(calc, 29),
        "ira45zUsdM": year_row(calc, 30),
        "selfDesignedUsdM": year_row(calc, 31),
        "totalUsdM": year_row(calc, 33),
        "discountFactor": year_row(calc, 34),
        "pvUsdM": year_row(calc, 35),
    }
    fossil = {
        "totalCapexUsdM": year_row(calc, 51),
        "totalOpexUsdM": year_row(calc, 52),
        "etsUsdM": year_row(calc, 54),
        "fuelEuUsdM": year_row(calc, 55),
        # The workbook has no fossil 45Z row (the credit is green-only); the
        # engine's uniform SideResult reports it as zero.
        "ira45zUsdM": [0] * HORIZON_YEARS,
        "selfDesignedUsdM": year_row(calc, 56),
        "totalUsdM": year_row(calc, 58),
        "discountFactor": year_row(calc, 59),
        "pvUsdM": year_row(calc, 60),
    }

    return {
        "source": {
            "workbook": WORKBOOK.name,
            "sha256": sha256(WORKBOOK),
            "basis": "cached workbook values (data_only=True)",
        },
        "summary": summary,
        "intermediates": intermediates,
        "perYear": {
            "green": green,
            "fossil": fossil,
            "co2AbatedTonnes": year_row(calc, 65),
        },
    }


def write_json(path: Path, data: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=1, ensure_ascii=False) + "\n", encoding="utf-8", newline="\n")


def main() -> None:
    wb_f = openpyxl.load_workbook(WORKBOOK, data_only=False)
    wb_v = openpyxl.load_workbook(WORKBOOK, data_only=True)
    assert wb_f.sheetnames == SHEETS, f"unexpected tabs: {wb_f.sheetnames}"

    write_json(DUMP_OUT, build_dump(wb_f, wb_v))
    write_json(EXPECTED_OUT, build_expected(wb_v))

    print(f"workbook sha256: {sha256(WORKBOOK)}")
    print(f"wrote {DUMP_OUT.relative_to(ROOT)}")
    print(f"wrote {EXPECTED_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
